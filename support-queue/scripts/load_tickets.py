#!/usr/bin/env python3
"""Нормализация выгрузки обращений (CSV/XLSX) в tickets.json.

    python3 load_tickets.py tickets.csv --out tickets.json
    python3 load_tickets.py export.xlsx --out tickets.json --map text=Комментарий

Колонки определяются по типовым названиям на русском и английском.
Найденный маппинг печатается — проверьте его прежде, чем идти дальше:
неверно угаданная колонка «текст» тихо превращает весь прогон в мусор.
"""

import argparse
import csv
import io
import json
import os
import re
import sys

ALIASES = {
    "id": ["id", "номер", "№", "ticket", "ticket_id", "тикет", "обращение", "case", "key"],
    "created": ["дата", "created", "created_at", "date", "время", "получено", "datetime", "дата создания"],
    "subject": ["тема", "subject", "title", "заголовок", "тема обращения"],
    "text": ["текст", "сообщение", "body", "description", "message", "comment", "комментарий",
             "вопрос", "содержание", "text", "обращение", "суть"],
    "customer": ["клиент", "автор", "from", "имя", "customer", "contact", "контакт", "фио", "name"],
    "email": ["email", "почта", "e-mail", "mail", "адрес"],
    "channel": ["канал", "channel", "source", "источник", "тип обращения"],
    "status": ["статус", "status", "state"],
    "priority": ["приоритет", "priority", "срочность"],
    "tier": ["тариф", "сегмент", "tier", "plan", "категория клиента"],
}


def norm(s):
    return re.sub(r"[\s_\-\.]+", " ", str(s or "").strip().lower())


def guess_map(headers):
    got, used = {}, set()
    nh = {h: norm(h) for h in headers}
    for field, aliases in ALIASES.items():
        best = None
        for h in headers:
            if h in used:
                continue
            v = nh[h]
            if v in aliases:
                best = h
                break
            if best is None and any(a in v for a in aliases):
                best = h
        if best:
            got[field] = best
            used.add(best)
    return got


def read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Для XLSX нужен openpyxl: pip install openpyxl --break-system-packages")
        ws = load_workbook(path, data_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sys.exit("Файл пуст")
        headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
        return headers, [dict(zip(headers, r)) for r in rows[1:]]

    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        sys.exit("Не удалось определить кодировку файла")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";" if sample.count(";") > sample.count(",") else ","
    rdr = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows = list(rdr)
    return (rdr.fieldnames or []), rows


def main():
    ap = argparse.ArgumentParser(description="Выгрузка обращений → tickets.json")
    ap.add_argument("input")
    ap.add_argument("--out", default="tickets.json")
    ap.add_argument("--map", action="append", default=[],
                    help="ручной маппинг, например --map text=Комментарий")
    ap.add_argument("--limit", type=int, help="взять только первые N (для пробного прогона)")
    a = ap.parse_args()

    headers, rows = read_rows(a.input)
    if not rows:
        sys.exit("В файле нет строк")

    mapping = guess_map(headers)
    for m in a.map:
        if "=" not in m:
            sys.exit(f"Неверный --map: {m}. Формат: поле=Колонка")
        k, v = m.split("=", 1)
        mapping[k.strip()] = v.strip()

    if "text" not in mapping:
        sys.exit("Не найдена колонка с текстом обращения. Задайте вручную: "
                 f"--map text=<название>\nДоступные колонки: {', '.join(map(str, headers))}")

    tickets = []
    for i, r in enumerate(rows, 1):
        def val(f):
            col = mapping.get(f)
            v = r.get(col) if col else None
            return None if v in (None, "") else str(v).strip()

        text = val("text")
        if not text:
            continue
        tickets.append({k: v for k, v in {
            "id": val("id") or f"T-{i:04d}",
            "created": val("created"),
            "subject": val("subject"),
            "text": text,
            "customer": val("customer"),
            "email": val("email"),
            "channel": val("channel"),
            "status": val("status"),
            "priority": val("priority"),
            "tier": val("tier"),
        }.items() if v is not None})
        if a.limit and len(tickets) >= a.limit:
            break

    with open(a.out, "w", encoding="utf-8") as f:
        json.dump({"source": os.path.basename(a.input), "tickets": tickets},
                  f, ensure_ascii=False, indent=2)

    print(f"Готово: {a.out} — обращений: {len(tickets)} из {len(rows)} строк")
    print("Маппинг колонок (проверьте!):")
    for k in ALIASES:
        if k in mapping:
            print(f"  {k:9} ← {mapping[k]}")
    missing = [k for k in ("id", "created", "subject", "customer") if k not in mapping]
    if missing:
        print(f"  не найдены: {', '.join(missing)} — не критично, но если они есть в файле,"
              " задайте через --map")
    skipped = len(rows) - len(tickets)
    if skipped > 0 and not a.limit:
        print(f"  пропущено строк без текста: {skipped}")


if __name__ == "__main__":
    main()
